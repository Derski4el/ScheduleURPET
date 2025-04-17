from fastapi import HTTPException
from sqlalchemy.orm import Session
import openpyxl
import models

def parse_schedule_excel(file, db: Session):
    """
    Парсит Excel-файл с расписанием и сохраняет данные в базу данных.
    Возвращает количество добавленных записей.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")
    
    # Чтение Excel-файла
    workbook = openpyxl.load_workbook(file.file)
    sheet = workbook.active
    schedules_added = 0

    # Чтение групп и аудиторий (строка 4: группы, строка 5: аудитории)
    groups_rooms = []
    for col in range(2, 53, 2):  # B4:AZ4 (столбцы 2, 4, ..., 52)
        group = sheet.cell(row=4, column=col).value
        room = sheet.cell(row=4, column=col + 1).value
        if group and room:
            groups_rooms.append((group, room))
    
    # Чтение дней недели (строка 5: B5:BA5)
    days = []
    for col in range(2, 54, 2):  # B5:BA5 (столбцы 2, 4, ..., 52)
        day_str = sheet.cell(row=5, column=col).value
        if day_str:
            days.append(day_str)
    
    # Проверка корректности данных
    if len(days) != 6:
        raise HTTPException(status_code=400, detail="Expected 6 days in row 5 (B5:BA5)")
    
    # Чтение времени и пар
    for day_idx, day_str in enumerate(days):
        # Каждая группа имеет 46 строк времени (A6:A52), делим на 6 дней (~7–8 пар на день)
        rows_per_day = 46 // 6  # Примерно 7–8 строк на день
        row_start = 6 + day_idx * rows_per_day
        row_end = row_start + rows_per_day
        
        for row in range(row_start, row_end):
            time_str = sheet.cell(row=row, column=1).value  # Время из A6:A52
            if not time_str:
                continue
            
            # Чтение пар для каждой группы
            for col_idx, (group_id, room) in enumerate(groups_rooms):
                col = 2 + col_idx * 2  # Столбцы B, D, F, ...
                pair_info = sheet.cell(row=row, column=col).value
                if not pair_info:
                    continue
                
                # Разделяем название пары и ФИО преподавателя
                try:
                    course_name, teacher_name = pair_info.split(", ")
                except ValueError:
                    raise HTTPException(status_code=400, detail=f"Invalid pair format at row {row}, column {col}: expected 'Course, Teacher'")
                
                # Поиск преподавателя в базе данных
                teacher = db.query(models.User).filter(models.User.name == teacher_name, models.User.role == "teacher").first()
                if not teacher:
                    raise HTTPException(status_code=400, detail=f"Teacher {teacher_name} not found in database")
                
                # Формируем course_id (например, Group1_Math_14апреля)
                course_id = f"{group_id}_{course_name}_{day_str.replace(' ', '_')}"
                
                # Формируем время (например, "14 апреля 08:00-09:30")
                full_time = f"{day_str.split(' ')[0]} {day_str.split(' ')[1]} {time_str}"
                
                # Сохраняем расписание в базу данных
                db_schedule = models.Schedule(
                    course_id=course_id,
                    time=full_time,
                    room=room,
                    teacher_id=teacher.id,
                    group_id=group_id
                )
                db.add(db_schedule)
                schedules_added += 1
    
    db.commit()
    return schedules_added